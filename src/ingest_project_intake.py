"""
Turn a filled-out project intake form into pipeline data.

Usage (from the repo root, after filling the yellow cells):
    python src/ingest_project_intake.py [path/to/form.xlsx]

Defaults to templates/project_intake.xlsx. The script reads the form,
validates it with the same fail-loud rules as every other intake, and
UPSERTS the project into data/projects/project_master.csv and
project_assumptions.csv (an existing project_id is replaced, with a
notice - nothing is ever silently duplicated). If validation fails,
the data files are left exactly as they were.

Then:  python src/build_project_appraisal.py  and refresh Page 6.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import pandas as pd

from financials.loader import ClientFSValidationError
from financials.projects import PROJECTS_DIR, load_projects
from financials.schemas import PROJECT_ASSUMPTION_CODES

DEFAULT_FORM = Path(__file__).resolve().parent.parent / "templates" / "project_intake.xlsx"

PROJECT_CELLS = {          # row -> master column (form rows 6..13, column B)
    6: "project_id", 7: "project_name", 8: "category",
    9: "initial_investment", 10: "start_period", 11: "horizon_years",
    12: "review_status", 13: "rationale",
}
LEVER_ROWS = range(17, 24)  # code col A, value col B, rationale col D


def parse_intake(path):
    """Read the form -> (master row dict, list of assumption row dicts)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit(
            "openpyxl is required to read the form: pip install openpyxl")
    ws = load_workbook(path, data_only=True)["Project Intake"]

    master = {}
    for row, column in PROJECT_CELLS.items():
        value = ws[f"B{row}"].value
        master[column] = "" if value is None else value
    missing = [c for c, v in master.items() if v == "" and c != "rationale"]
    if missing:
        raise SystemExit(
            f"form incomplete - fill the yellow cells for: {missing}")
    if master["rationale"] == "":
        master["rationale"] = "(no rationale given)"

    assumptions = []
    for row in LEVER_ROWS:
        code = ws[f"A{row}"].value
        value = ws[f"B{row}"].value
        if code is None or value is None or str(value).strip() == "":
            continue
        if code not in PROJECT_ASSUMPTION_CODES:
            raise SystemExit(f"unknown assumption_code on row {row}: {code}")
        unit = ws[f"C{row}"].value or ""
        rationale = ws[f"D{row}"].value or "(no rationale given)"
        assumptions.append({
            "project_id": master["project_id"],
            "assumption_code": code,
            "value": float(value),
            "unit": unit,
            "rationale": rationale,
        })
    if not assumptions:
        raise SystemExit("no levers filled in - a project must change "
                         "something (revenue and/or savings)")
    return master, assumptions


def upsert(master_row, assumption_rows, projects_dir=None):
    """Replace-or-append the project; validate; roll back on failure."""
    projects_dir = Path(projects_dir) if projects_dir else PROJECTS_DIR
    master_path = projects_dir / "project_master.csv"
    assum_path = projects_dir / "project_assumptions.csv"
    master_before = master_path.read_text(encoding="utf-8")
    assum_before = assum_path.read_text(encoding="utf-8")

    master = pd.read_csv(master_path, dtype=str)
    assumptions = pd.read_csv(assum_path, dtype=str)
    pid = str(master_row["project_id"])
    replacing = pid in set(master["project_id"])
    master = master[master["project_id"] != pid]
    assumptions = assumptions[assumptions["project_id"] != pid]
    master = pd.concat(
        [master, pd.DataFrame([master_row])[master.columns.tolist()]],
        ignore_index=True)
    new_rows = pd.DataFrame(assumption_rows)[assumptions.columns.tolist()]
    assumptions = pd.concat([assumptions, new_rows], ignore_index=True)
    master.to_csv(master_path, index=False)
    assumptions.to_csv(assum_path, index=False)

    try:
        load_projects(projects_dir=projects_dir, strict=True)
    except ClientFSValidationError:
        master_path.write_text(master_before, encoding="utf-8")
        assum_path.write_text(assum_before, encoding="utf-8")
        raise
    return replacing


def main():
    form = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FORM
    if not form.exists():
        raise SystemExit(f"form not found: {form}")
    master_row, assumption_rows = parse_intake(form)
    try:
        replacing = upsert(master_row, assumption_rows)
    except ClientFSValidationError as exc:
        print()
        print("INTAKE REJECTED (data files unchanged):")
        print(exc)
        raise SystemExit(1)

    verb = "replaced" if replacing else "added"
    print()
    print(f"project {master_row['project_id']} {verb}: "
          f"{master_row['project_name']} "
          f"(${master_row['initial_investment']}M over "
          f"{master_row['horizon_years']}y, {len(assumption_rows)} levers)")
    print()
    print("next:  python src/build_project_appraisal.py   then Refresh "
          "the report (Page 6)")


if __name__ == "__main__":
    main()
